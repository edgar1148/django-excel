import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def calculate_tax(row):
    """
    Рассчитывает налог на основе значения в столбце 'Налоговая база'.

    param:
    - row (pandas.Series): Строка DataFrame.
    return:
    - float: Рассчитанный налог.
    """
    tax_base = row['Unnamed: 4']
    
    if pd.isna(tax_base):
        return 0
    
    tax_rate = 0.13 if tax_base <= 5000000 else 0.15
    return round(tax_base * tax_rate)

def calculate_deviation(row):
    """
    Рассчитывает отклонение между значениями
    в столбце 'Исчислено всего' и 'Исчислено всего по формуле'.

    param:
    - row (pandas.Series): Строка DataFrame.
    return:
    - float: Рассчитанное отклонение.
    """

    if pd.isna(row['Unnamed: 1']):
        return 0
    tax_column = pd.to_numeric(row['Исчислено всего'], errors='coerce')
    calculated_column = row['Исчислено всего по формуле']
    
    if pd.notna(tax_column):
        return tax_column - calculated_column
    else:
        return 0

def process_excel(input_file, output_file):
    """
    Обрабатывает файл Excel, добавляет новые
    столбцы и сохраняет результат в новом
    файле Excel в соответствии с требованиями.

    param:
    - input_file (str): Путь к входному файлу Excel.
    - output_file (str): Путь для сохранения нового файла Excel.

    return:
    - None
    """

    try:
        df = pd.read_excel(input_file, skipfooter=1, header=1)
    except Exception as e:
        print(f"Ошибка при чтении файла Excel: {e}")
        return
    
    column_names = df.columns.tolist()
    print(column_names)


    


    df['Исчислено всего по формуле'] = df.apply(calculate_tax, axis=1)
    df['Отклонения'] = df.apply(calculate_deviation, axis=1)
    df['Филиал'] = df['Unnamed: 0']
    df['Сотрудник'] = df['Unnamed: 1']
    df['Налоговая база'] = df['Unnamed: 4']

    df = df[df['Сотрудник'].notna()]
    
    df = df.sort_values(by='Отклонения', ascending=False)

    new_df = df[['Филиал', 'Сотрудник', 'Налоговая база', 'Исчислено всего', 'Исчислено всего по формуле', 'Отклонения']]
    
    new_df.to_excel(output_file, index=False)

    wb = Workbook()
    ws = wb.active

    header_data = [('D1', 'Налог'), ('A1', 'Филиал'), ('B1', 'Сотрудник'), ('C1', 'Налоговая база'), ('F1', 'Отклонения')]
    column_data = [('D2', 'Исчислено всего'), ('E2', 'Исчислено всего по формуле')]

    ws.merge_cells('D1:E1')

    for cell, value in header_data:
        ws[cell] = value
        ws[cell].alignment = Alignment(horizontal='center')

    for cell, value in column_data:
        ws[cell] = value
        ws[cell].alignment = Alignment(horizontal='center')

    for col_range in ['A', 'B', 'C', 'F']:
        ws.merge_cells(f'{col_range}1:{col_range}2')
        ws[col_range + '1'].alignment = Alignment(horizontal='center')

    

    for row in dataframe_to_rows(new_df, index=False, header=False):
        ws.append(row)

    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    green_fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')

    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        cell = row[-1]
        if cell.value == 0:
            cell.fill = green_fill
        else:
            cell.fill = red_fill
            cell.font = Font(color='FFFFFF')

    wb.save(output_file)
